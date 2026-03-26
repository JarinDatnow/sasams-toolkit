"""
=============================================================================
  BALMORAL TOP 10 — FULL PIPELINE
  Query SASAMS → Generate Top 10 spreadsheets → Print them
  
  One click. One script. Done.
=============================================================================
  Setup:
    1. pip install pyodbc openpyxl pywin32
    2. Copy config.example.py to config.py and fill in your details
    3. Put TEMPLATE__top_10.xlsx in the templates/ folder
    4. Run:  python top10_pipeline.py

  Flags:
    --no-print     Generate spreadsheets only, don't print
    --discover     Show database tables and columns
=============================================================================
"""

import os
import sys
import pyodbc
from pathlib import Path
from collections import defaultdict

# ── Load config ─────────────────────────────────────────────────────────────
try:
    from config import DB_PATH, DB_PASSWORD, DATA_YEAR, TERM, OUTPUT_FOLDER
except ImportError:
    print("[FAIL] config.py not found.")
    print("       Copy config.example.py to config.py and fill in your values.")
    sys.exit(1)

TEMPLATE_NAME = "TEMPLATE__top_10.xlsx"
AUTO_PRINT = True


def connect():
    driver = "{Microsoft Access Driver (*.mdb, *.accdb)}"
    conn_str = f"DRIVER={driver};DBQ={os.path.abspath(DB_PATH)};PWD={DB_PASSWORD};"
    try:
        conn = pyodbc.connect(conn_str)
        print("[OK] Connected to database")
        return conn
    except (pyodbc.Error, UnicodeDecodeError) as e:
        print(f"[FAIL] Could not connect: {e}")
        sys.exit(1)


def discover_schema(conn):
    cursor = conn.cursor()
    print("\n" + "=" * 60)
    print("  DATABASE SCHEMA DISCOVERY")
    print("=" * 60)
    tables_of_interest = ["Learner_Info", "LearnerPromotion", "Classes"]
    for table in cursor.tables(tableType='TABLE'):
        tname = table.table_name
        if tname.startswith("MSys"):
            continue
        marker = "  <<<" if tname in tables_of_interest else ""
        print(f"\n  TABLE: {tname}{marker}")
        for col in cursor.columns(table=tname):
            marker2 = "  <<<" if tname in tables_of_interest else ""
            print(f"    {col.column_name:30s} {col.type_name}{marker2}")
    print("\n" + "=" * 60)
    print("Check these tables for field names if the query ever needs tweaking.")
    print("=" * 60)


def query_top10(conn):
    sql = f"""
    SELECT 
        P.Grade, 
        C.ClassName,
        L.SName, 
        L.FName, 
        P.LearnerAverage,
        P.DataYear
    FROM (Learner_Info L 
    INNER JOIN LearnerPromotion P ON L.ID = P.LearnerId)
    LEFT JOIN Classes C ON L.Class = C.ClassId
    WHERE P.DataYear = '{DATA_YEAR}' 
      AND P.LearnerAverage IS NOT NULL
      AND P.LearnerAverage > 0
      AND P.ReportId = (
        SELECT MAX(P2.ReportId) 
        FROM LearnerPromotion P2 
        WHERE P2.LearnerId = P.LearnerId 
        AND P2.DataYear = '{DATA_YEAR}'
      )
    ORDER BY P.Grade, C.ClassName, P.LearnerAverage DESC
    """

    cursor = conn.cursor()
    try:
        cursor.execute(sql)
        rows = cursor.fetchall()
        print(f"[OK] Query returned {len(rows)} learners")
    except (pyodbc.Error, UnicodeDecodeError) as e:
        print(f"[FAIL] Query error: {e}")
        sys.exit(1)

    PER_GRADE_GRADES = {'7', '8', '9', '10', '11', '12', '07', '08', '09'}
    groups = defaultdict(list)

    for row in rows:
        grade, classname, surname, firstname, average, datayear = row
        grade_str = str(grade).strip() if grade else ""
        class_str = str(classname).strip() if classname else ""

        if grade_str in PER_GRADE_GRADES:
            group_key = grade_str.lstrip('0') or '0'
        else:
            group_key = class_str if class_str else f"Grade{grade_str}"

        groups[group_key].append({
            'surname': str(surname).upper() if surname else '',
            'firstname': str(firstname).upper() if firstname else '',
            'grade': grade_str,
            'average': round(float(average)) if average else 0,
        })

    top10 = {}
    for key, learners in groups.items():
        top10[key] = learners[:10]

    print(f"[OK] Found {len(top10)} groups: {', '.join(sorted(top10.keys()))}")
    return top10


def make_grade_label(group_key):
    k = group_key.strip()
    if k in ('0', '00', 'R'):
        k = 'R'
    elif k.startswith('0') and k[1:].isdigit():
        k = k.lstrip('0')
    return f"GRADE  {k}"


def generate_xlsx(top10_groups, template_path, output_dir):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment

    TITLE_FONT  = Font(name='Calibri', bold=True, size=20, color='FF000000')
    HEADER_FONT = Font(name='Calibri', bold=True, size=16, color='FF000000')
    NR_FONT     = Font(name='Calibri', bold=True, size=22, color='FF000000')
    DATA_FONT   = Font(name='Calibri', bold=True, size=18, color='FF000000')
    CENTER      = Alignment(horizontal='center')

    output_dir.mkdir(exist_ok=True)
    files_created = []

    for group_key in sorted(top10_groups.keys()):
        learners = top10_groups[group_key]
        grade_label = make_grade_label(group_key)

        wb = load_workbook(str(template_path))
        ws = wb['Worksheet']

        ws['C2'] = 'TOP 10';       ws['C2'].font = TITLE_FONT
        ws['C4'] = grade_label;     ws['C4'].font = TITLE_FONT
        ws['C6'] = TERM;            ws['C6'].font = TITLE_FONT

        ws['A11'] = 'Nr';           ws['A11'].font = HEADER_FONT
        ws['B11'] = 'SURNAME';      ws['B11'].font = HEADER_FONT; ws['B11'].alignment = CENTER
        ws['C11'] = ' NAME';        ws['C11'].font = HEADER_FONT; ws['C11'].alignment = CENTER
        ws['D11'] = '%';            ws['D11'].font = HEADER_FONT; ws['D11'].alignment = CENTER

        for i, l in enumerate(learners):
            row = 13 + i
            ws.cell(row=row, column=1, value=i + 1).font = NR_FONT
            ws.cell(row=row, column=2, value=l['surname']).font = DATA_FONT
            ws.cell(row=row, column=3, value=l['firstname']).font = DATA_FONT
            ws.cell(row=row, column=4, value=l['average']).font = DATA_FONT

        for i in range(len(learners), 10):
            row = 13 + i
            for col in range(1, 5):
                ws.cell(row=row, column=col, value=None)

        out_name = f"{group_key}_top10.xlsx"
        out_path = output_dir / out_name
        wb.save(str(out_path))
        files_created.append(out_path)

        top = learners[0] if learners else {}
        print(f"  {out_name:20s} {grade_label:15s} #1 {top.get('surname','')} {top.get('average','')}%")

    return files_created


def mass_print(files):
    try:
        import win32com.client
    except ImportError:
        print("[SKIP] pywin32 not installed — can't auto-print. Run: pip install pywin32")
        return

    print(f"\n[>>] Printing {len(files)} files to default printer...")
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False

    printed = 0
    try:
        for f in files:
            try:
                wb = excel.Workbooks.Open(str(f.resolve()))
                ws = wb.Worksheets("Worksheet")
                ws.PrintOut()
                wb.Close(SaveChanges=False)
                printed += 1
            except Exception as e:
                print(f"  ERROR printing {f.name}: {e}")
        print(f"[OK] Sent {printed}/{len(files)} to printer")
    finally:
        excel.Quit()


def main():
    args = sys.argv[1:]
    script_dir = Path(__file__).parent

    if '--discover' in args:
        conn = connect()
        discover_schema(conn)
        conn.close()
        return

    no_print = '--no-print' in args or not AUTO_PRINT

    print("=" * 60)
    print("  BALMORAL TOP 10 PIPELINE")
    print(f"  Year: {DATA_YEAR}  |  {TERM}")
    print("=" * 60)

    template_path = script_dir / "templates" / TEMPLATE_NAME
    if not template_path.exists():
        # Also check script directory directly
        template_path = script_dir / TEMPLATE_NAME
    if not template_path.exists():
        print(f"[FAIL] Template not found. Put {TEMPLATE_NAME} in the templates/ folder.")
        sys.exit(1)

    print("\n[STEP 1] Querying SASAMS database...")
    conn = connect()
    top10 = query_top10(conn)
    conn.close()

    print(f"\n[STEP 2] Generating spreadsheets...")
    output_dir = script_dir / OUTPUT_FOLDER
    files = generate_xlsx(top10, template_path, output_dir)
    print(f"[OK] Created {len(files)} files in {OUTPUT_FOLDER}/")

    if no_print:
        print(f"\n[STEP 3] Printing skipped (--no-print)")
    else:
        print(f"\n[STEP 3] Printing...")
        mass_print(files)

    print("\n" + "=" * 60)
    print(f"  DONE — {len(files)} top 10 lists generated")
    if not no_print:
        print(f"  Sent to printer")
    print(f"  Files in: {output_dir}")
    print("=" * 60)


if __name__ == '__main__':
    main()
