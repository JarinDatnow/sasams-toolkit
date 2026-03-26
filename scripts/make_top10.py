"""
TOP 10 GENERATOR (D6 Export Version)
=====================================
Use this if you export top achiever reports from D6 instead of querying SASAMS directly.

1. Put TEMPLATE__top_10.xlsx in this folder
2. Copy all your D6 export .xlsx files into this folder
3. Run:  python make_top10.py
4. Output files land in a "top10_output" subfolder

Auto-detects grade label and PERCENTAGE column from each file.
"""

import os
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment
except ImportError:
    print("openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

TEMPLATE_NAME = "TEMPLATE__top_10.xlsx"
OUTPUT_FOLDER = "top10_output"

TITLE_FONT  = Font(name='Calibri', bold=True, size=20, color='FF000000')
HEADER_FONT = Font(name='Calibri', bold=True, size=16, color='FF000000')
NR_FONT     = Font(name='Calibri', bold=True, size=22, color='FF000000')
DATA_FONT   = Font(name='Calibri', bold=True, size=18, color='FF000000')
CENTER      = Alignment(horizontal='center')


def find_pct_column(ws):
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=11, column=c).value
        if v and 'PERCENTAGE' in str(v).upper():
            return c
    for c in range(ws.max_column, 0, -1):
        if ws.cell(row=13, column=c).value is not None:
            return c
    return None


def get_grade_label(ws):
    raw = ws['D4'].value
    if raw:
        return 'GRADE  ' + raw.replace('GRADE', '').strip()
    return 'UNKNOWN'


def get_term(ws):
    raw = ws['D6'].value
    if raw:
        t = raw.replace('TERMS:', '').replace('TERM:', '').strip()
        return f'TERM {t}'
    return 'TERM 1'


def process_file(src_path, template_path, output_path):
    src = openpyxl.load_workbook(src_path)
    src_ws = src['Worksheet']

    grade_label = get_grade_label(src_ws)
    term_str = get_term(src_ws)
    pct_col = find_pct_column(src_ws)

    if pct_col is None:
        print(f"  !! Could not find PERCENTAGE column — skipping")
        return False

    data = []
    for r in range(13, 23):
        nr = src_ws.cell(row=r, column=1).value
        learner = src_ws.cell(row=r, column=2).value
        surname = src_ws.cell(row=r, column=3).value
        name = src_ws.cell(row=r, column=4).value
        pct = src_ws.cell(row=r, column=pct_col).value
        if nr is not None:
            if isinstance(learner, (int, float)):
                learner = str(int(learner))
            data.append((nr, learner, surname, name, pct))

    if not data:
        print(f"  !! No data found — skipping")
        return False

    wb = openpyxl.load_workbook(template_path)
    ws = wb['Worksheet']

    ws['C2'] = 'TOP 10';       ws['C2'].font = TITLE_FONT
    ws['C4'] = grade_label;     ws['C4'].font = TITLE_FONT
    ws['C6'] = term_str;        ws['C6'].font = TITLE_FONT

    ws['A11'] = 'Nr';           ws['A11'].font = HEADER_FONT
    ws['B11'] = 'SURNAME';      ws['B11'].font = HEADER_FONT; ws['B11'].alignment = CENTER
    ws['C11'] = ' NAME';        ws['C11'].font = HEADER_FONT; ws['C11'].alignment = CENTER
    ws['D11'] = '%';            ws['D11'].font = HEADER_FONT; ws['D11'].alignment = CENTER

    for i, (nr, learner, surname, name, pct) in enumerate(data):
        row = 13 + i
        ws.cell(row=row, column=1, value=nr).font = NR_FONT
        ws.cell(row=row, column=2, value=surname).font = DATA_FONT
        ws.cell(row=row, column=3, value=name).font = DATA_FONT
        ws.cell(row=row, column=4, value=pct).font = DATA_FONT

    for i in range(len(data), 10):
        row = 13 + i
        for col in range(1, 5):
            ws.cell(row=row, column=col, value=None)

    wb.save(output_path)
    print(f"  -> {grade_label} | {term_str} | {len(data)} learners | #1 {data[0][2]} {data[0][4]}%")
    return True


def main():
    folder = Path('.')
    template = folder / TEMPLATE_NAME

    if not template.exists():
        print(f"ERROR: Put {TEMPLATE_NAME} in this folder and try again.")
        sys.exit(1)

    source_files = sorted([
        f for f in folder.glob('*.xlsx')
        if f.name != TEMPLATE_NAME and '_top10' not in f.name
    ])

    if not source_files:
        print("ERROR: No source .xlsx files found. Copy them into this folder.")
        sys.exit(1)

    out_dir = folder / OUTPUT_FOLDER
    out_dir.mkdir(exist_ok=True)

    print(f"Template: {TEMPLATE_NAME}")
    print(f"Found {len(source_files)} source file(s)")
    print(f"Output:  {OUTPUT_FOLDER}/")
    print("-" * 50)

    done = 0
    for src in source_files:
        out_name = src.stem + '_top10.xlsx'
        out_path = out_dir / out_name
        print(f"{src.name}:")
        if process_file(str(src), str(template), str(out_path)):
            done += 1

    print("-" * 50)
    print(f"Done! {done}/{len(source_files)} files created in {OUTPUT_FOLDER}/")


if __name__ == '__main__':
    main()
